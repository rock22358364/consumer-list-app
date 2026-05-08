import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font
from copy import copy
import re
import os
import zipfile
import time
from xml.etree import ElementTree as ET

# =================================================
# 固定配置
# =================================================
SOURCE_SHEET = "Equipment List"
TEMPLATE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "consumer_template.xlsx")
DATA_START_ROW = 7

SRC_COLS = {
    "pid": 1,
    "industrial_complex": 2,
    "process_area": 3,
    "subprocess_area": 4,
    "tag_no": 5,
    "equipment_name": 6,
    "velocity": 19,
    "power": 20,
    "voltage": 21,
    "inquiry_group": 41,
    "vfd": 49,
}

# =================================================
# 多语言文本
# =================================================
TEXTS = {
    "zh": {
        "title": "⚡ Electrical Consumer List Generator",
        "subtitle": "从设备清单（Equipment List）中自动提取电气用电设备信息，生成标准化的 Electrical Consumer List。",
        "rules_title": "📋 使用规则",
        "rule_1": "源文件基于 [EQP-APAC-TPL-1101-Equipment List](https://covestro.sharepoint.com/:x:/r/sites/002331/_layouts/15/Doc.aspx?sourcedoc=%7BA9768CC5-50DA-419E-BBA0-C9D5BDDBE2CC%7D&file=EQP-APAC-TPL-1101-Equipment%20List.xlsx&action=default&mobileredirect=true&DefaultItemOpen=1%3Fweb%3D1) 模板",
        "rule_2": "筛选逻辑：功率（Power）或电压（Voltage）列中任意一列有有效数值 → 该行纳入 Consumer List",
        "rule_3": "以下内容不视为有效数值：`-`、空值、`NA`、`N/A`",
        "steps_title": "🚀 使用步骤",
        "step_1": "上传 Equipment List 文件（.xlsx）",
        "step_2": "系统自动识别电气设备数量，展开 Preview 可预览",
        "step_3": "点击 **Generate Consumer List**",
        "step_4": "点击 **Download** 下载结果文件",
        "notes_title": "⚠️ 注意事项",
        "note_1": "上传文件不会被服务器保存，关闭页面即清除",
        "note_2": "建议生成后打开文件核对数据准确性",
        "upload_label": "上传 Equipment List 文件（.xlsx）",
        "reading_file": "正在读取文件...",
        "processing": "正在处理数据...",
        "generating": "正在生成 Consumer List...",
        "error_no_data": "❌ 无法找到数据起始行。",
        "vfd_found": "🔄 VFD 列已识别，位于源文件第 {} 列",
        "no_vfd": "ℹ️ 源文件中未找到 VFD 列",
        "motor_detected": "✅ 识别到 **{}** 个电气用电设备（共 {} 行数据）",
        "vfd_detected": "🔄 其中 **{}** 个设备带 VFD",
        "strikethrough_detected": "📝 检测到 {} 个带删除线的单元格",
        "preview_title": "📋 数据预览",
        "btn_generate": "🚀 生成 Consumer List",
        "error_no_sheet": "❌ 模板中未找到 Consumer List 工作表。",
        "warning_missing_cols": "以下模板列未匹配到（将跳过）：{}",
        "done": "✅ 完成！已写入 **{}** 行数据。",
        "btn_download": "📥 下载 Consumer List",
        "upload_hint": "👆 请上传 Equipment List 文件开始使用",
    },
    "en": {
        "title": "⚡ Electrical Consumer List Generator",
        "subtitle": "Automatically extracts electrical consumer information from the Equipment List and generates a standardized Electrical Consumer List.",
        "rules_title": "📋 Rules",
        "rule_1": "Source file must be based on the [EQP-APAC-TPL-1101-Equipment List](https://covestro.sharepoint.com/:x:/r/sites/002331/_layouts/15/Doc.aspx?sourcedoc=%7BA9768CC5-50DA-419E-BBA0-C9D5BDDBE2CC%7D&file=EQP-APAC-TPL-1101-Equipment%20List.xlsx&action=default&mobileredirect=true&DefaultItemOpen=1%3Fweb%3D1) template",
        "rule_2": "Selection logic: Any row with a valid numeric value in either the Power or Voltage column will be included",
        "rule_3": "The following are NOT considered valid values: `-`, blank, `NA`, `N/A`",
        "steps_title": "🚀 How to Use",
        "step_1": "Upload the Equipment List file (.xlsx)",
        "step_2": "The system will automatically identify electrical consumers; expand Preview to verify",
        "step_3": "Click **Generate Consumer List**",
        "step_4": "Click **Download** to save the output file",
        "notes_title": "⚠️ Notes",
        "note_1": "Uploaded files are NOT stored on the server; they are cleared once the page is closed",
        "note_2": "It is recommended to review the generated file for accuracy after download",
        "upload_label": "Upload Equipment List (.xlsx)",
        "reading_file": "Reading file...",
        "processing": "Processing data...",
        "generating": "Generating Consumer List...",
        "error_no_data": "❌ Cannot find data start row.",
        "vfd_found": "🔄 VFD column found at source column {}",
        "no_vfd": "ℹ️ No VFD column found in source file.",
        "motor_detected": "✅ **{}** motor consumers detected (from {} total rows)",
        "vfd_detected": "🔄 **{}** equipment with VFD detected",
        "strikethrough_detected": "📝 Detected {} cells with strikethrough formatting",
        "preview_title": "📋 Preview Data",
        "btn_generate": "🚀 Generate Consumer List",
        "error_no_sheet": "❌ Cannot find Consumer List sheet in template.",
        "warning_missing_cols": "Could not match these template columns (will skip): {}",
        "done": "✅ Done! **{}** rows written to Consumer List.",
        "btn_download": "📥 Download Consumer List",
        "upload_hint": "👆 Please upload an Equipment List file to get started",
    },
}

# =================================================
# 页面配置
# =================================================
st.set_page_config(
    page_title="Electrical Consumer List Generator",
    page_icon="⚡",
    layout="centered",
)

# 语言切换
if "lang" not in st.session_state:
    st.session_state.lang = "zh"

col_title, col_lang = st.columns([5, 1])
with col_lang:
    lang_options = {"zh": "中文", "en": "EN"}
    selected_lang = st.selectbox(
        "🌐",
        options=list(lang_options.keys()),
        format_func=lambda x: lang_options[x],
        index=0 if st.session_state.lang == "zh" else 1,
        label_visibility="collapsed",
    )
    if selected_lang != st.session_state.lang:
        st.session_state.lang = selected_lang
        st.rerun()

T = TEXTS[st.session_state.lang]

with col_title:
    st.title(T["title"])

st.markdown(T["subtitle"])

# 功能说明折叠区
with st.expander(T["rules_title"], expanded=False):
    st.markdown(f"1. {T['rule_1']}")
    st.markdown(f"2. {T['rule_2']}")
    st.markdown(f"3. {T['rule_3']}")

with st.expander(T["steps_title"], expanded=False):
    st.markdown(f"1. {T['step_1']}")
    st.markdown(f"2. {T['step_2']}")
    st.markdown(f"3. {T['step_3']}")
    st.markdown(f"4. {T['step_4']}")

with st.expander(T["notes_title"], expanded=False):
    st.markdown(f"- {T['note_1']}")
    st.markdown(f"- {T['note_2']}")

st.divider()

uploaded_file = st.file_uploader(T["upload_label"], type=["xlsx"])


# =================================================
# 工具函数
# =================================================
def norm(x):
    if x is None:
        return ""
    return re.sub(r"\s+", "", str(x).lower())


def find_consumer_columns(ws):
    col_text = {}
    for c in range(1, ws.max_column + 1):
        texts = []
        for r in range(1, 7):
            v = ws.cell(r, c).value
            if v:
                texts.append(str(v))
        col_text[c] = norm(" ".join(texts))

    rules = {
        "industrial_complex": "complex",
        "process_area": "processarea",
        "subprocess_area": "subprocess",
        "tag_no": "tagno",
        "equipment_id": "设备编号",
        "equipment_name": "servicedescription",
        "pid": "p&id",
        "inquiry_group": "packageno",
        "voltage": "voltage(v)",
        "power_installed": "installedpower",
        "power_rated": "ratedpower",
        "velocity": "ratedspeed",
        "vfd": "vfd",
    }

    result = {}
    used_cols = set()
    for field, pattern in rules.items():
        for col, joined in col_text.items():
            if col in used_cols:
                continue
            if pattern in joined:
                result[field] = col
                used_cols.add(col)
                break
    return result


def find_vfd_source_col(raw, data_start):
    for col_idx in range(raw.shape[1]):
        for row_idx in range(min(data_start, 10)):
            val = raw.iloc[row_idx, col_idx]
            if val is not None and not pd.isna(val):
                s = str(val).strip().upper()
                if s == "VFD" or "变频" in str(val):
                    return col_idx
    return None


def copy_row_style(ws, source_row, target_row, max_col):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for c in range(1, max_col + 1):
        src_cell = ws.cell(source_row, c)
        tgt_cell = ws.cell(target_row, c)
        if src_cell.font:
            tgt_cell.font = copy(src_cell.font)
        if src_cell.border:
            tgt_cell.border = copy(src_cell.border)
        if src_cell.alignment:
            tgt_cell.alignment = copy(src_cell.alignment)
        if src_cell.fill:
            tgt_cell.fill = copy(src_cell.fill)
        if src_cell.number_format:
            tgt_cell.number_format = src_cell.number_format


def unmerge_data_area(ws, start_row):
    merges_to_remove = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= start_row:
            merges_to_remove.append(str(merged_range))
    for m in merges_to_remove:
        ws.unmerge_cells(m)


def build_equipment_id(row):
    parts = []
    for k in ["industrial_complex", "process_area", "subprocess_area", "tag_no"]:
        v = row.get(k)
        if v and str(v).strip() and str(v).strip().lower() != "nan":
            parts.append(str(v).strip())
    return "-".join(parts)


def to_number(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "-", "--", "—", "/", "N/A", "n/a", "NA", "nan"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def is_valid_motor_value(val):
    return to_number(val) is not None


def col_idx_to_letter(col_idx):
    result = ""
    col_idx += 1
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result


def get_excel_cell_ref(excel_row, col_idx_0based):
    return f"{col_idx_to_letter(col_idx_0based)}{excel_row}"


def get_vfd_value(val):
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "nan", "-", "NA", "N/A"):
        return None
    return s


def read_strikethrough_from_xlsx(file_obj):
    strikethrough_cells = set()
    try:
        file_obj.seek(0)
        zf = zipfile.ZipFile(file_obj)

        ns = {
            "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        wb_xml = ET.fromstring(zf.read("xl/workbook.xml"))
        sheet_rid = None
        for sheet_elem in wb_xml.findall(".//main:sheet", ns):
            if sheet_elem.get("name") == SOURCE_SHEET:
                sheet_rid = sheet_elem.get(
                    "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
                )
                break

        if not sheet_rid:
            zf.close()
            return strikethrough_cells

        rels_xml = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        sheet_path = None
        for rel in rels_xml.findall(
            ".//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship"
        ):
            if rel.get("Id") == sheet_rid:
                target = rel.get("Target")
                sheet_path = f"xl/{target}" if not target.startswith("/") else target.lstrip("/")
                break

        if not sheet_path:
            zf.close()
            return strikethrough_cells

        styles_xml = ET.fromstring(zf.read("xl/styles.xml"))
        fonts_elem = styles_xml.find(".//main:fonts", ns)
        strikethrough_font_indices = set()

        if fonts_elem is not None:
            for font_idx, font_elem in enumerate(fonts_elem.findall("main:font", ns)):
                strike_elem = font_elem.find("main:strike", ns)
                if strike_elem is not None:
                    val_attr = strike_elem.get("val", "1")
                    if val_attr != "0":
                        strikethrough_font_indices.add(font_idx)

        cell_xfs = styles_xml.find(".//main:cellXfs", ns)
        strikethrough_xf_indices = set()
        if cell_xfs is not None:
            for xf_idx, xf_elem in enumerate(cell_xfs.findall("main:xf", ns)):
                font_id = xf_elem.get("fontId")
                if font_id is not None and int(font_id) in strikethrough_font_indices:
                    strikethrough_xf_indices.add(xf_idx)

        sheet_xml = ET.fromstring(zf.read(sheet_path))
        for row_elem in sheet_xml.findall(".//main:row", ns):
            for cell_elem in row_elem.findall("main:c", ns):
                style_idx = cell_elem.get("s")
                if style_idx is not None and int(style_idx) in strikethrough_xf_indices:
                    cell_ref = cell_elem.get("r")
                    if cell_ref:
                        strikethrough_cells.add(cell_ref)

        zf.close()
    except Exception:
        pass

    return strikethrough_cells


# =================================================
# 主逻辑
# =================================================
if uploaded_file:

    # ===== 读取文件（带进度条） =====
    progress_bar = st.progress(0, text=T["reading_file"])
    raw = pd.read_excel(
        uploaded_file,
        sheet_name=SOURCE_SHEET,
        header=None,
        engine="calamine",
    )
    progress_bar.progress(30, text=T["processing"])

    # 找到数据起始行
    data_start = None
    for i in range(raw.shape[0]):
        first_val = raw.iloc[i, 0]
        if first_val is not None and str(first_val).strip() == "1":
            second_val = raw.iloc[i, 1]
            if second_val is not None and str(second_val).strip() not in ("", "nan"):
                data_start = i
                break

    if data_start is None:
        progress_bar.empty()
        st.error(T["error_no_data"])
        st.stop()

    # 动态查找 VFD 列
    vfd_col_idx = find_vfd_source_col(raw, data_start)
    if vfd_col_idx is not None:
        SRC_COLS["vfd"] = vfd_col_idx

    progress_bar.progress(50, text=T["processing"])

    # 提取数据
    df = raw.iloc[data_start:].reset_index(drop=True)
    df = df.dropna(how="all").reset_index(drop=True)
    df = df[df.iloc[:, 0].notna()].reset_index(drop=True)

    motor_data = []
    for idx in range(len(df)):
        row_data = {}
        for field, col_idx in SRC_COLS.items():
            if col_idx < len(df.columns):
                val = df.iloc[idx, col_idx]
                row_data[field] = None if pd.isna(val) else val
            else:
                row_data[field] = None
        row_data["_df_idx"] = idx
        motor_data.append(row_data)

    filtered_data = []
    for row in motor_data:
        if is_valid_motor_value(row.get("power")) or is_valid_motor_value(row.get("voltage")):
            row["power_num"] = to_number(row.get("power"))
            row["voltage_num"] = to_number(row.get("voltage"))
            row["velocity_num"] = to_number(row.get("velocity"))
            row["vfd_value"] = get_vfd_value(row.get("vfd"))
            filtered_data.append(row)

    progress_bar.progress(70, text=T["processing"])

    # 读取删除线信息
    strikethrough_cells = read_strikethrough_from_xlsx(uploaded_file)

    progress_bar.progress(100, text="✅")
    time.sleep(0.5)
    progress_bar.empty()

    # ===== 显示结果 =====
    if vfd_col_idx is not None:
        st.info(T["vfd_found"].format(col_idx_to_letter(vfd_col_idx)))
    else:
        st.info(T["no_vfd"])

    vfd_count = sum(1 for r in filtered_data if r.get("vfd_value"))
    st.success(T["motor_detected"].format(len(filtered_data), len(df)))
    if vfd_count > 0:
        st.info(T["vfd_detected"].format(vfd_count))
    if strikethrough_cells:
        st.info(T["strikethrough_detected"].format(len(strikethrough_cells)))

    with st.expander(T["preview_title"]):
        preview_df = pd.DataFrame([
            {
                "No.": i + 1,
                "Tag No.": r.get("tag_no"),
                "Equipment Name": r.get("equipment_name"),
                "Power (kW)": r.get("power_num"),
                "Voltage (V)": r.get("voltage_num"),
                "Speed (rpm)": r.get("velocity_num"),
                "VFD": r.get("vfd_value") or "",
            }
            for i, r in enumerate(filtered_data)
        ])
        st.dataframe(preview_df, use_container_width=True, hide_index=True)

    st.divider()

    if st.button(T["btn_generate"], type="primary", use_container_width=True):

        # ===== 生成文件（带进度条） =====
        gen_progress = st.progress(0, text=T["generating"])

        wb = load_workbook(TEMPLATE_FILE)

        target_ws_name = None
        for name in wb.sheetnames:
            if "consumer" in name.lower() and "electrical" in name.lower():
                target_ws_name = name
                break
        if target_ws_name is None:
            for name in wb.sheetnames:
                if "consumer" in name.lower():
                    target_ws_name = name
                    break

        if target_ws_name is None:
            gen_progress.empty()
            st.error(T["error_no_sheet"])
            st.stop()

        ws = wb[target_ws_name]
        unmerge_data_area(ws, DATA_START_ROW)
        col_map = find_consumer_columns(ws)

        missing_cols = set(["industrial_complex", "process_area", "subprocess_area",
                           "tag_no", "equipment_id", "equipment_name", "pid",
                           "inquiry_group", "voltage", "power_installed",
                           "power_rated", "velocity", "vfd"]) - set(col_map)
        if missing_cols:
            st.warning(T["warning_missing_cols"].format(missing_cols))

        gen_progress.progress(10, text=T["generating"])

        max_col = ws.max_column
        excel_data_start_row = data_start + 1
        r = DATA_START_ROW
        sn = 1
        total_rows = len(filtered_data)

        for row_idx, row_data in enumerate(filtered_data):
            equip_id = build_equipment_id(row_data)

            if r > DATA_START_ROW:
                copy_row_style(ws, DATA_START_ROW, r, max_col)

            ws.cell(r, 1).value = sn

            src_excel_row = excel_data_start_row + row_data["_df_idx"]

            field_values = {
                "industrial_complex": (row_data.get("industrial_complex"), SRC_COLS["industrial_complex"]),
                "process_area": (row_data.get("process_area"), SRC_COLS["process_area"]),
                "subprocess_area": (row_data.get("subprocess_area"), SRC_COLS["subprocess_area"]),
                "tag_no": (row_data.get("tag_no"), SRC_COLS["tag_no"]),
                "equipment_id": (equip_id, None),
                "equipment_name": (row_data.get("equipment_name"), SRC_COLS["equipment_name"]),
                "pid": (row_data.get("pid"), SRC_COLS["pid"]),
                "inquiry_group": (row_data.get("inquiry_group"), SRC_COLS["inquiry_group"]),
                "voltage": (row_data.get("voltage_num"), SRC_COLS["voltage"]),
                "power_installed": (row_data.get("power_num"), SRC_COLS["power"]),
                "power_rated": (row_data.get("power_num"), SRC_COLS["power"]),
                "velocity": (row_data.get("velocity_num"), SRC_COLS["velocity"]),
                "vfd": (row_data.get("vfd_value"), SRC_COLS.get("vfd")),
            }

            for field, col_idx in col_map.items():
                if field not in field_values:
                    continue

                val, src_col_idx = field_values[field]

                if val is None:
                    continue
                if isinstance(val, float) and pd.isna(val):
                    continue

                cell = ws.cell(r, col_idx)
                cell.value = val

                if strikethrough_cells and src_col_idx is not None:
                    cell_ref = get_excel_cell_ref(src_excel_row, src_col_idx)
                    if cell_ref in strikethrough_cells:
                        existing_font = cell.font
                        cell.font = Font(
                            name=existing_font.name,
                            size=existing_font.size,
                            bold=existing_font.bold,
                            italic=existing_font.italic,
                            color=existing_font.color,
                            strikethrough=True,
                        )

            r += 1
            sn += 1

            # 更新进度条
            if total_rows > 0:
                progress_pct = int(10 + (row_idx + 1) / total_rows * 80)
                gen_progress.progress(progress_pct, text=T["generating"])

        # 保存文件
        gen_progress.progress(95, text=T["generating"])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        gen_progress.progress(100, text="✅")
        time.sleep(0.5)
        gen_progress.empty()

        st.success(T["done"].format(sn - 1))

        st.download_button(
            T["btn_download"],
            data=buf,
            file_name="Consumer_List_Generated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

else:
    st.info(T["upload_hint"])
